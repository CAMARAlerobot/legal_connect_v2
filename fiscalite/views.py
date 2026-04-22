from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count
from decimal import Decimal, InvalidOperation
from datetime import date, timedelta
import json

from .models import Declaration, Echeance, CommentaireDeclaration, TYPES_IMPOT
from . import calculateur

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY


# ── Helpers ────────────────────────────────────────────────────────

MOIS_NOMS = ['','Janvier','Février','Mars','Avril','Mai','Juin',
             'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

def _periode_label(d):
    label = f"{d.get_periode_display()} {d.annee}"
    if d.mois:
        label += f" — {MOIS_NOMS[d.mois]}"
    if d.trimestre:
        label += f" — T{d.trimestre}"
    return label


# ── Dashboard ──────────────────────────────────────────────────────

@login_required
def dashboard_fiscal(request):
    user         = request.user
    declarations = Declaration.objects.filter(utilisateur=user).order_by('-created_at')
    echeances    = Echeance.objects.filter(utilisateur=user).order_by('date_limite')
    aujourd_hui  = timezone.now().date()

    echeances_urgentes = echeances.filter(
        date_limite__gte=aujourd_hui,
        date_limite__lte=aujourd_hui + timedelta(days=7),
        statut='a_faire'
    )
    echeances.filter(date_limite__lt=aujourd_hui, statut='a_faire').update(statut='en_retard')

    # Données graphique par type d'impôt
    par_type = (
        declarations
        .values('type_impot')
        .annotate(total=Sum('montant_impot'), nb=Count('id'))
        .order_by('-total')
    )
    chart_labels = [dict(TYPES_IMPOT).get(x['type_impot'], x['type_impot']) for x in par_type]
    chart_data   = [float(x['total'] or 0) for x in par_type]

    # Données graphique évolution mensuelle (12 derniers mois)
    annee_courante = aujourd_hui.year
    evol_labels, evol_data = [], []
    for m in range(1, aujourd_hui.month + 1):
        total_m = declarations.filter(annee=annee_courante, mois=m).aggregate(s=Sum('montant_impot'))['s'] or 0
        evol_labels.append(MOIS_NOMS[m][:3])
        evol_data.append(float(total_m))

    context = {
        'declarations'       : declarations[:5],
        'echeances'          : echeances[:8],
        'echeances_urgentes' : echeances_urgentes.count(),
        'total_declarations' : declarations.count(),
        'total_impots'       : declarations.aggregate(s=Sum('montant_impot'))['s'] or 0,
        'aujourd_hui'        : aujourd_hui,
        'chart_labels_json'  : json.dumps(chart_labels),
        'chart_data_json'    : json.dumps(chart_data),
        'evol_labels_json'   : json.dumps(evol_labels),
        'evol_data_json'     : json.dumps(evol_data),
        'declarations_en_revision': declarations.filter(soumise_a_expert=True, statut='en_revision').count(),
    }
    return render(request, 'fiscalite/dashboard.html', context)


# ── Calculateur ────────────────────────────────────────────────────

@login_required
def calculateur_view(request):
    resultat    = None
    type_impot  = request.POST.get('type_impot', 'tva')
    erreur      = None
    suggestions = []

    if request.method == 'POST':
        try:
            data = {
                'chiffre_affaires': request.POST.get('chiffre_affaires', 0) or 0,
                'charges'         : request.POST.get('charges', 0) or 0,
                'salaire'         : request.POST.get('salaire', 0) or 0,
                'nb_employes'     : request.POST.get('nb_employes', 1) or 1,
            }
            resultat    = calculateur.calculer(type_impot, data)
            suggestions = calculateur.suggestions_optimisation(type_impot, data)

            if request.POST.get('sauvegarder') and resultat:
                periode = request.POST.get('periode', 'mensuel')
                mois_val = timezone.now().month if periode == 'mensuel' else None
                trimestre_val = ((timezone.now().month - 1) // 3) + 1 if periode == 'trimestriel' else None
                Declaration.objects.create(
                    utilisateur      = request.user,
                    type_impot       = type_impot,
                    periode          = periode,
                    annee            = timezone.now().year,
                    mois             = mois_val,
                    trimestre        = trimestre_val,
                    chiffre_affaires = Decimal(str(data['chiffre_affaires'])),
                    charges          = Decimal(str(data['charges'])),
                    benefice_net     = resultat.get('benefice_net', 0),
                    taux_applique    = resultat.get('taux', 0) if isinstance(resultat.get('taux'), (int, float)) else 0,
                    montant_impot    = resultat.get('montant_impot', 0),
                )
                messages.success(request, 'Déclaration sauvegardée dans votre historique !')
                return redirect('fiscalite:historique')

        except (ValueError, InvalidOperation):
            erreur = "Erreur de calcul : veuillez vérifier vos saisies."

    context = {
        'types_impot': TYPES_IMPOT,
        'type_actif' : type_impot,
        'resultat'   : resultat,
        'erreur'     : erreur,
        'post_data'  : request.POST if request.method == 'POST' else {},
        'suggestions': suggestions,
        'periodes'   : [('mensuel','Mensuel'), ('trimestriel','Trimestriel'), ('annuel','Annuel')],
    }
    return render(request, 'fiscalite/calculateur.html', context)


# ── Calculateur de pénalités ───────────────────────────────────────

@login_required
def penalites_view(request):
    resultat = None
    erreur   = None

    if request.method == 'POST':
        try:
            montant        = Decimal(request.POST.get('montant', 0) or 0)
            date_echeance  = date.fromisoformat(request.POST.get('date_echeance', ''))
            date_paiement  = date.fromisoformat(request.POST.get('date_paiement', '')) \
                             if request.POST.get('date_paiement') else date.today()
            resultat = calculateur.calculer_penalites(montant, date_echeance, date_paiement)
        except (ValueError, InvalidOperation):
            erreur = "Veuillez vérifier les valeurs saisies."

    return render(request, 'fiscalite/penalites.html', {
        'resultat'    : resultat,
        'erreur'      : erreur,
        'post_data'   : request.POST if request.method == 'POST' else {},
        'aujourd_hui' : date.today().isoformat(),
    })


# ── Historique ─────────────────────────────────────────────────────

@login_required
def historique(request):
    declarations = Declaration.objects.filter(utilisateur=request.user)

    type_impot = request.GET.get('type', '')
    annee      = request.GET.get('annee', '')
    statut     = request.GET.get('statut', '')
    if type_impot:
        declarations = declarations.filter(type_impot=type_impot)
    if annee:
        declarations = declarations.filter(annee=annee)
    if statut:
        declarations = declarations.filter(statut=statut)

    total_impots = declarations.aggregate(s=Sum('montant_impot'))['s'] or 0
    annees = Declaration.objects.filter(
        utilisateur=request.user
    ).values_list('annee', flat=True).distinct()

    context = {
        'declarations': declarations,
        'types_impot' : TYPES_IMPOT,
        'type_actif'  : type_impot,
        'annee_active': annee,
        'statut_actif': statut,
        'annees'      : sorted(set(annees), reverse=True),
        'total_impots': total_impots,
        'statuts'     : [('en_attente','En attente'),('soumise','Soumise'),
                         ('en_revision','En révision'),('validee','Validée'),('en_retard','En retard')],
    }
    return render(request, 'fiscalite/historique.html', context)


# ── Export Excel ───────────────────────────────────────────────────

@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    declarations = Declaration.objects.filter(utilisateur=request.user).order_by('-annee', '-created_at')

    type_impot = request.GET.get('type', '')
    annee      = request.GET.get('annee', '')
    if type_impot:
        declarations = declarations.filter(type_impot=type_impot)
    if annee:
        declarations = declarations.filter(annee=annee)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Déclarations Fiscales"

    BLEU   = "1B4F72"
    BLEU_C = "D6EAF8"

    entete_font    = Font(bold=True, color="FFFFFF", size=11)
    entete_fill    = PatternFill("solid", fgColor=BLEU)
    entete_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    titre_font     = Font(bold=True, color=BLEU, size=14)
    sous_titre_font = Font(italic=True, color="555555", size=10)
    total_fill     = PatternFill("solid", fgColor=BLEU_C)
    total_font     = Font(bold=True, color=BLEU, size=11)
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = "LÉGAL CONNECT — Déclarations Fiscales"
    ws['A1'].font = titre_font
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Exporté le {date.today().strftime('%d/%m/%Y')} — Côte d'Ivoire"
    ws['A2'].font = sous_titre_font
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.append([])  # ligne vide

    colonnes = [
        "Type d'impôt", "Période", "Année", "Mois",
        "CA / Salaire (FCFA)", "Charges (FCFA)", "Bénéfice net (FCFA)",
        "Montant impôt (FCFA)", "Statut", "Date création"
    ]
    ws.append(colonnes)
    row_entete = ws.max_row
    for col_idx, _ in enumerate(colonnes, start=1):
        cell = ws.cell(row=row_entete, column=col_idx)
        cell.font   = entete_font
        cell.fill   = entete_fill
        cell.alignment = entete_align
        cell.border = thin_border

    # Données
    alternating = PatternFill("solid", fgColor="F0F7FF")
    for i, d in enumerate(declarations):
        row = [
            d.get_type_impot_display(),
            d.get_periode_display(),
            d.annee,
            MOIS_NOMS[d.mois] if d.mois else ('T' + str(d.trimestre) if d.trimestre else '—'),
            float(d.chiffre_affaires),
            float(d.charges),
            float(d.benefice_net),
            float(d.montant_impot),
            d.get_statut_display(),
            d.created_at.strftime('%d/%m/%Y'),
        ]
        ws.append(row)
        r = ws.max_row
        fill = alternating if i % 2 == 0 else None
        for col_idx in range(1, len(colonnes) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if col_idx in (5, 6, 7, 8):
                cell.number_format = '#,##0'

    # Ligne total
    ws.append([])
    total_row = ["TOTAL", "", "", "", "", "", "",
                 float(declarations.aggregate(s=Sum('montant_impot'))['s'] or 0), "", ""]
    ws.append(total_row)
    r = ws.max_row
    for col_idx in range(1, len(colonnes) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill   = total_fill
        cell.font   = total_font
        cell.border = thin_border
        if col_idx == 8:
            cell.number_format = '#,##0'

    # Largeurs colonnes
    largeurs = [30, 15, 8, 12, 22, 18, 22, 22, 15, 14]
    for i, larg in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[row_entete].height = 30

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nom = f"declarations_fiscales_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    wb.save(response)
    return response


# ── Récapitulatif annuel ───────────────────────────────────────────

@login_required
def recapitulatif_annuel(request):
    annee_courante = timezone.now().year
    annee = int(request.GET.get('annee', annee_courante))

    declarations = Declaration.objects.filter(utilisateur=request.user, annee=annee)
    annees_dispo  = Declaration.objects.filter(
        utilisateur=request.user
    ).values_list('annee', flat=True).distinct()

    # Par type d'impôt
    par_type = []
    for code, label in TYPES_IMPOT:
        decls = declarations.filter(type_impot=code)
        total = decls.aggregate(s=Sum('montant_impot'))['s'] or 0
        if total > 0 or decls.exists():
            par_type.append({
                'code' : code,
                'label': label,
                'nb'   : decls.count(),
                'total': total,
                'decls': decls,
            })

    # Par mois
    par_mois = []
    for m in range(1, 13):
        total_m = declarations.filter(mois=m).aggregate(s=Sum('montant_impot'))['s'] or 0
        par_mois.append({'mois': MOIS_NOMS[m], 'num': m, 'total': total_m})

    total_annuel = declarations.aggregate(s=Sum('montant_impot'))['s'] or 0

    context = {
        'annee'       : annee,
        'annees_dispo': sorted(set(annees_dispo), reverse=True),
        'declarations': declarations,
        'par_type'    : par_type,
        'par_mois'    : par_mois,
        'total_annuel': total_annuel,
        'nb_total'    : declarations.count(),
        'chart_mois_labels': json.dumps([MOIS_NOMS[m][:3] for m in range(1, 13)]),
        'chart_mois_data'  : json.dumps([float(declarations.filter(mois=m).aggregate(s=Sum('montant_impot'))['s'] or 0) for m in range(1, 13)]),
    }
    return render(request, 'fiscalite/recapitulatif.html', context)


# ── Statistiques avancées ──────────────────────────────────────────

@login_required
def statistiques(request):
    user  = request.user
    decls = Declaration.objects.filter(utilisateur=user)
    ech   = Echeance.objects.filter(utilisateur=user)
    annee_courante = timezone.now().year

    # Évolution sur 3 ans
    annees_graph = [annee_courante - 2, annee_courante - 1, annee_courante]
    evol_par_annee = []
    for a in annees_graph:
        total = decls.filter(annee=a).aggregate(s=Sum('montant_impot'))['s'] or 0
        evol_par_annee.append(float(total))

    # Répartition par type
    repartition = []
    for code, label in TYPES_IMPOT:
        total = decls.filter(type_impot=code).aggregate(s=Sum('montant_impot'))['s'] or 0
        if float(total) > 0:
            repartition.append({'label': label.split('—')[0].strip(), 'total': float(total)})

    # Taux de conformité (déclarations validées vs total)
    nb_total    = decls.count()
    nb_validees = decls.filter(statut='validee').count()
    nb_retard   = decls.filter(statut='en_retard').count()
    taux_conformite = (nb_validees / nb_total * 100) if nb_total > 0 else 0

    # Échéances stats
    ech_faites    = ech.filter(statut='fait').count()
    ech_en_retard = ech.filter(statut='en_retard').count()
    ech_a_faire   = ech.filter(statut='a_faire').count()

    # Évolution mensuelle année courante
    evol_mois = []
    for m in range(1, 13):
        total_m = decls.filter(annee=annee_courante, mois=m).aggregate(s=Sum('montant_impot'))['s'] or 0
        evol_mois.append(float(total_m))

    context = {
        'nb_total'       : nb_total,
        'nb_validees'    : nb_validees,
        'nb_retard'      : nb_retard,
        'taux_conformite': round(taux_conformite, 1),
        'total_impots'   : decls.aggregate(s=Sum('montant_impot'))['s'] or 0,
        'ech_faites'     : ech_faites,
        'ech_en_retard'  : ech_en_retard,
        'ech_a_faire'    : ech_a_faire,
        'annees_graph'   : json.dumps(annees_graph),
        'evol_annee_data': json.dumps(evol_par_annee),
        'repartition_labels': json.dumps([r['label'] for r in repartition]),
        'repartition_data'  : json.dumps([r['total'] for r in repartition]),
        'evol_mois_labels'  : json.dumps([MOIS_NOMS[m][:3] for m in range(1, 13)]),
        'evol_mois_data'    : json.dumps(evol_mois),
        'annee_courante' : annee_courante,
    }
    return render(request, 'fiscalite/statistiques.html', context)


# ── Rapport de conformité ──────────────────────────────────────────

@login_required
def rapport_conformite(request):
    user  = request.user
    decls = Declaration.objects.filter(utilisateur=user)
    ech   = Echeance.objects.filter(utilisateur=user)
    aujourd_hui = date.today()

    # Points de contrôle
    checks = []

    # 1. Déclarations en retard
    en_retard = decls.filter(statut='en_retard')
    checks.append({
        'titre'   : 'Déclarations en retard',
        'ok'      : en_retard.count() == 0,
        'detail'  : f"{en_retard.count()} déclaration(s) en retard" if en_retard.exists() else "Aucune déclaration en retard",
        'items'   : list(en_retard[:5]),
        'priorite': 'haute',
    })

    # 2. Échéances dépassées non soldées
    ech_retard = ech.filter(statut='en_retard')
    checks.append({
        'titre'   : 'Échéances fiscales dépassées',
        'ok'      : ech_retard.count() == 0,
        'detail'  : f"{ech_retard.count()} échéance(s) dépassée(s)" if ech_retard.exists() else "Toutes les échéances sont à jour",
        'items'   : list(ech_retard[:5]),
        'priorite': 'haute',
    })

    # 3. Déclarations TVA ce mois-ci
    mois_courant = aujourd_hui.month
    annee_courante = aujourd_hui.year
    tva_ce_mois = decls.filter(type_impot='tva', annee=annee_courante, mois=mois_courant)
    checks.append({
        'titre'   : f'Déclaration TVA — {MOIS_NOMS[mois_courant]} {annee_courante}',
        'ok'      : tva_ce_mois.exists(),
        'detail'  : "TVA déclarée pour ce mois" if tva_ce_mois.exists() else "Aucune déclaration TVA ce mois-ci",
        'items'   : [],
        'priorite': 'moyenne',
    })

    # 4. IS annuel
    is_annuel = decls.filter(type_impot='is', annee=annee_courante, periode='annuel')
    checks.append({
        'titre'   : f'Déclaration IS annuelle {annee_courante}',
        'ok'      : is_annuel.exists(),
        'detail'  : "IS annuel déclaré" if is_annuel.exists() else "IS annuel non encore déclaré pour cette année",
        'items'   : [],
        'priorite': 'haute',
    })

    # 5. CNPS
    cnps_ce_mois = decls.filter(type_impot='cnps', annee=annee_courante, mois=mois_courant)
    checks.append({
        'titre'   : f'Cotisations CNPS — {MOIS_NOMS[mois_courant]} {annee_courante}',
        'ok'      : cnps_ce_mois.exists(),
        'detail'  : "CNPS déclaré pour ce mois" if cnps_ce_mois.exists() else "CNPS non déclaré ce mois-ci",
        'items'   : [],
        'priorite': 'moyenne',
    })

    # 6. Déclarations soumises à expert en attente
    en_revision = decls.filter(soumise_a_expert=True, statut='en_revision')
    checks.append({
        'titre'   : 'Déclarations en cours de révision expert',
        'ok'      : True,
        'detail'  : f"{en_revision.count()} déclaration(s) en cours de révision" if en_revision.exists() else "Aucune révision en cours",
        'items'   : list(en_revision[:3]),
        'priorite': 'info',
    })

    nb_ok  = sum(1 for c in checks if c['ok'])
    score  = round(nb_ok / len(checks) * 100)

    context = {
        'checks'      : checks,
        'nb_ok'       : nb_ok,
        'nb_total'    : len(checks),
        'score'       : score,
        'aujourd_hui' : aujourd_hui,
    }
    return render(request, 'fiscalite/conformite.html', context)


# ── Workflow expert fiscal ─────────────────────────────────────────

@login_required
def soumettre_expert(request, pk):
    decl = get_object_or_404(Declaration, pk=pk, utilisateur=request.user)
    if request.method == 'POST':
        decl.soumise_a_expert      = True
        decl.statut                = 'en_revision'
        decl.date_soumission_expert = timezone.now()
        decl.save()
        messages.success(request, f'Déclaration #{decl.pk} soumise pour révision expert.')
    return redirect('fiscalite:detail_declaration', pk=pk)


@login_required
def detail_declaration(request, pk):
    decl         = get_object_or_404(Declaration, pk=pk, utilisateur=request.user)
    commentaires = decl.commentaires.select_related('auteur').all()
    return render(request, 'fiscalite/detail_declaration.html', {
        'decl'        : decl,
        'commentaires': commentaires,
        'periode_label': _periode_label(decl),
    })


@login_required
def ajouter_commentaire(request, pk):
    decl = get_object_or_404(Declaration, pk=pk, utilisateur=request.user)
    if request.method == 'POST':
        texte = request.POST.get('texte', '').strip()
        if texte:
            est_expert = hasattr(request.user, 'profil') and request.user.profil.specialite == 'droit_fiscal'
            CommentaireDeclaration.objects.create(
                declaration=decl,
                auteur=request.user,
                texte=texte,
                est_expert=est_expert,
            )
            messages.success(request, 'Commentaire ajouté.')
        else:
            messages.error(request, 'Le commentaire ne peut pas être vide.')
    return redirect('fiscalite:detail_declaration', pk=pk)


# ── Calendrier ─────────────────────────────────────────────────────

@login_required
def calendrier(request):
    aujourd_hui = timezone.now().date()
    echeances   = Echeance.objects.filter(utilisateur=request.user).order_by('date_limite')

    # Vue mensuelle
    annee_cal = int(request.GET.get('annee', aujourd_hui.year))
    mois_cal  = int(request.GET.get('mois',  aujourd_hui.month))

    import calendar
    cal = calendar.monthcalendar(annee_cal, mois_cal)
    ech_du_mois = echeances.filter(date_limite__year=annee_cal, date_limite__month=mois_cal)
    ech_par_jour = {}
    for e in ech_du_mois:
        ech_par_jour.setdefault(e.date_limite.day, []).append(e)

    # Navigation mois
    if mois_cal == 1:
        mois_prec, annee_prec = 12, annee_cal - 1
    else:
        mois_prec, annee_prec = mois_cal - 1, annee_cal
    if mois_cal == 12:
        mois_suiv, annee_suiv = 1, annee_cal + 1
    else:
        mois_suiv, annee_suiv = mois_cal + 1, annee_cal

    context = {
        'echeances'   : echeances,
        'types_impot' : TYPES_IMPOT,
        'aujourd_hui' : aujourd_hui,
        'cal_semaines': cal,
        'annee_cal'   : annee_cal,
        'mois_cal'    : mois_cal,
        'mois_nom'    : MOIS_NOMS[mois_cal],
        'ech_par_jour': ech_par_jour,
        'mois_prec'   : mois_prec,
        'annee_prec'  : annee_prec,
        'mois_suiv'   : mois_suiv,
        'annee_suiv'  : annee_suiv,
        'jours_semaine': ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim'],
    }
    return render(request, 'fiscalite/calendrier.html', context)


@login_required
def ajouter_echeance(request):
    if request.method == 'POST':
        titre       = request.POST.get('titre')
        type_impot  = request.POST.get('type_impot')
        date_limite = request.POST.get('date_limite')
        montant     = request.POST.get('montant') or None
        notes       = request.POST.get('notes', '')

        if titre and type_impot and date_limite:
            Echeance.objects.create(
                utilisateur = request.user,
                titre       = titre,
                type_impot  = type_impot,
                date_limite = date_limite,
                montant     = montant,
                notes       = notes,
            )
            messages.success(request, 'Échéance ajoutée au calendrier !')
        else:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
    return redirect('fiscalite:calendrier')


@login_required
def marquer_echeance(request, pk):
    ech = get_object_or_404(Echeance, pk=pk, utilisateur=request.user)
    ech.statut = 'fait'
    ech.save()
    messages.success(request, f'Échéance "{ech.titre}" marquée comme faite.')
    return redirect('fiscalite:calendrier')


@login_required
def supprimer_echeance(request, pk):
    ech = get_object_or_404(Echeance, pk=pk, utilisateur=request.user)
    if request.method == 'POST':
        ech.delete()
        messages.success(request, 'Échéance supprimée.')
    return redirect('fiscalite:calendrier')


# ── Export PDF déclaration ─────────────────────────────────────────

@login_required
def exporter_pdf_declaration(request, pk):
    decl = get_object_or_404(Declaration, pk=pk, utilisateur=request.user)
    user = request.user

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="declaration_{decl.type_impot}_{decl.annee}_{decl.pk}.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    BLEU    = colors.HexColor('#1B4F72')
    BLEU_S  = colors.HexColor('#2E86C1')
    BLEU_CL = colors.HexColor('#D6EAF8')
    VERT    = colors.HexColor('#1E8449')
    VERT_CL = colors.HexColor('#EAFAF1')

    style_titre     = ParagraphStyle('titre', fontSize=18, textColor=BLEU, alignment=TA_CENTER, spaceAfter=4, fontName='Helvetica-Bold')
    style_sous_titre = ParagraphStyle('sous_titre', fontSize=11, textColor=BLEU_S, alignment=TA_CENTER, spaceAfter=16, fontName='Helvetica')
    style_section   = ParagraphStyle('section', fontSize=11, textColor=BLEU, fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=6)
    style_normal    = ParagraphStyle('normal', fontSize=10, leading=16, alignment=TA_JUSTIFY, spaceAfter=6, fontName='Helvetica')
    style_footer    = ParagraphStyle('footer', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

    el = []

    el.append(Paragraph("LÉGAL CONNECT", style_titre))
    el.append(Paragraph("Plateforme Numérique d'Assistance Juridique", style_sous_titre))
    el.append(HRFlowable(width="100%", thickness=2, color=BLEU))
    el.append(Spacer(1, 0.3*cm))
    el.append(Paragraph("DÉCLARATION FISCALE", style_titre))
    el.append(Paragraph(decl.get_type_impot_display(), style_sous_titre))

    ref_table = Table([[
        f"Référence : #{decl.pk:04d}",
        f"Type : {decl.get_type_impot_display()}",
        f"Statut : {decl.get_statut_display()}",
        f"Date : {decl.created_at.strftime('%d/%m/%Y')}",
    ]], colWidths=[3.5*cm, 4.5*cm, 4*cm, 4.5*cm])
    ref_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), BLEU_CL),
        ('TEXTCOLOR',  (0,0), (-1,-1), BLEU),
        ('FONTNAME',   (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('ROWHEIGHT',  (0,0), (-1,-1), 22),
        ('BOX',        (0,0), (-1,-1), 0.5, BLEU_S),
    ]))
    el.append(ref_table)
    el.append(Spacer(1, 0.4*cm))

    el.append(HRFlowable(width="100%", thickness=1, color=BLEU_S))
    el.append(Paragraph("INFORMATIONS DU CONTRIBUABLE", style_section))
    try:
        profil = user.profil
        entreprise = profil.entreprise or 'Non renseignée'
        telephone  = profil.telephone  or 'Non renseigné'
        adresse    = profil.adresse    or 'Non renseignée'
    except Exception:
        entreprise = telephone = adresse = 'Non renseigné'

    contrib_data = [
        ['Nom complet', f"{user.first_name} {user.last_name}"],
        ['Email',       user.email or 'Non renseigné'],
        ['Entreprise',  entreprise],
        ['Téléphone',   telephone],
        ['Adresse',     adresse],
    ]
    contrib_table = Table(contrib_data, colWidths=[4*cm, 12.5*cm])
    contrib_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), BLEU),
        ('TEXTCOLOR',  (0,0), (0,-1), colors.white),
        ('FONTNAME',   (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',   (1,0), (1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWHEIGHT',  (0,0), (-1,-1), 20),
        ('LEFTPADDING',(0,0), (-1,-1), 8),
        ('GRID',       (0,0), (-1,-1), 0.5, BLEU_S),
        ('ROWBACKGROUNDS',(0,0),(-1,-1), [colors.HexColor('#F8FBFF'), colors.white]),
    ]))
    el.append(contrib_table)
    el.append(Spacer(1, 0.3*cm))

    el.append(HRFlowable(width="100%", thickness=1, color=BLEU_S))
    el.append(Paragraph("PÉRIODE DE DÉCLARATION", style_section))
    el.append(Paragraph(_periode_label(decl), style_normal))
    el.append(Spacer(1, 0.2*cm))

    el.append(HRFlowable(width="100%", thickness=1, color=BLEU_S))
    el.append(Paragraph("DONNÉES FINANCIÈRES", style_section))

    fin_data = [['Libellé', 'Montant (FCFA)']]
    if decl.chiffre_affaires:
        fin_data.append(["Chiffre d'affaires", f"{decl.chiffre_affaires:,.0f}"])
    if decl.charges:
        fin_data.append(["Charges déductibles", f"{decl.charges:,.0f}"])
    if decl.benefice_net:
        fin_data.append(["Bénéfice net", f"{decl.benefice_net:,.0f}"])
    if decl.taux_applique:
        fin_data.append(["Taux appliqué", f"{decl.taux_applique}%"])

    fin_table = Table(fin_data, colWidths=[10*cm, 6.5*cm])
    fin_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), BLEU),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,0), (-1,-1), 9),
        ('ALIGN',        (1,0), (1,-1), 'RIGHT'),
        ('ROWHEIGHT',    (0,0), (-1,-1), 22),
        ('LEFTPADDING',  (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('GRID',         (0,0), (-1,-1), 0.5, BLEU_S),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.HexColor('#F8FBFF'), colors.white]),
    ]))
    el.append(fin_table)
    el.append(Spacer(1, 0.4*cm))

    el.append(HRFlowable(width="100%", thickness=1, color=VERT))
    montant_table = Table([
        ["MONTANT DE L'IMPÔT DÛ"],
        [f"{decl.montant_impot:,.0f} FCFA"],
    ], colWidths=[16.5*cm])
    montant_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), BLEU),
        ('BACKGROUND', (0,1), (-1,1), VERT_CL),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('TEXTCOLOR',  (0,1), (-1,1), VERT),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 11),
        ('FONTSIZE',   (0,1), (-1,1), 22),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('ROWHEIGHT',  (0,0), (-1,0), 28),
        ('ROWHEIGHT',  (0,1), (-1,1), 50),
        ('BOX',        (0,0), (-1,-1), 1.5, VERT),
    ]))
    el.append(montant_table)
    el.append(Spacer(1, 0.5*cm))

    if decl.notes:
        el.append(HRFlowable(width="100%", thickness=1, color=BLEU_S))
        el.append(Paragraph("NOTES", style_section))
        el.append(Paragraph(decl.notes, style_normal))

    el.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    el.append(Spacer(1, 0.2*cm))
    el.append(Paragraph(
        f"Document généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} — "
        "Légal Connect — Université Nangui Abrogoua — Master 1 MIAGE 2025-2026 — Confidentiel",
        style_footer
    ))

    doc.build(el)
    return response


# ── Export PDF récapitulatif annuel ───────────────────────────────

@login_required
def export_pdf_recapitulatif(request):
    annee = int(request.GET.get('annee', timezone.now().year))
    user  = request.user
    declarations = Declaration.objects.filter(utilisateur=user, annee=annee)
    total_annuel = declarations.aggregate(s=Sum('montant_impot'))['s'] or 0

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="recapitulatif_fiscal_{annee}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    BLEU   = colors.HexColor('#1B4F72')
    BLEU_S = colors.HexColor('#2E86C1')
    VERT   = colors.HexColor('#1E8449')
    VERT_C = colors.HexColor('#EAFAF1')

    s_titre    = ParagraphStyle('t', fontSize=16, textColor=BLEU, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=4)
    s_sous     = ParagraphStyle('s', fontSize=10, textColor=BLEU_S, alignment=TA_CENTER, fontName='Helvetica', spaceAfter=12)
    s_section  = ParagraphStyle('sec', fontSize=11, textColor=BLEU, fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=5)
    s_footer   = ParagraphStyle('f', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)

    el = []
    el.append(Paragraph("LÉGAL CONNECT", s_titre))
    el.append(Paragraph("Récapitulatif Fiscal Annuel", s_sous))
    el.append(Paragraph(f"Exercice {annee}", s_sous))
    el.append(HRFlowable(width="100%", thickness=2, color=BLEU))
    el.append(Spacer(1, 0.3*cm))

    # Tableau récap par type
    el.append(Paragraph("RÉPARTITION PAR TYPE D'IMPÔT", s_section))
    rows = [['Type d\'impôt', 'Nb déclarations', 'Total (FCFA)']]
    for code, label in TYPES_IMPOT:
        decls_type = declarations.filter(type_impot=code)
        total_type = decls_type.aggregate(s=Sum('montant_impot'))['s'] or 0
        if decls_type.exists():
            rows.append([label, str(decls_type.count()), f"{float(total_type):,.0f}"])
    rows.append(['TOTAL', str(declarations.count()), f"{float(total_annuel):,.0f}"])

    t = Table(rows, colWidths=[9*cm, 4*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), BLEU),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',     (0,1), (-1,-2), 'Helvetica'),
        ('FONTNAME',     (0,-1),(-1,-1),'Helvetica-Bold'),
        ('BACKGROUND',   (0,-1),(-1,-1), VERT_C),
        ('TEXTCOLOR',    (0,-1),(-1,-1), VERT),
        ('FONTSIZE',     (0,0), (-1,-1), 9),
        ('ALIGN',        (1,0), (-1,-1), 'CENTER'),
        ('GRID',         (0,0), (-1,-1), 0.5, BLEU_S),
        ('ROWBACKGROUNDS',(0,1),(-1,-2), [colors.HexColor('#F8FBFF'), colors.white]),
        ('ROWHEIGHT',    (0,0), (-1,-1), 22),
    ]))
    el.append(t)

    # Total encadré
    el.append(Spacer(1, 0.5*cm))
    total_t = Table([["TOTAL IMPÔTS PAYÉS EN " + str(annee)], [f"{float(total_annuel):,.0f} FCFA"]], colWidths=[16.5*cm])
    total_t.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0), BLEU),
        ('BACKGROUND', (0,1),(-1,1), VERT_C),
        ('TEXTCOLOR',  (0,0),(-1,0), colors.white),
        ('TEXTCOLOR',  (0,1),(-1,1), VERT),
        ('FONTNAME',   (0,0),(-1,-1),'Helvetica-Bold'),
        ('FONTSIZE',   (0,0),(-1,0), 11),
        ('FONTSIZE',   (0,1),(-1,1), 20),
        ('ALIGN',      (0,0),(-1,-1),'CENTER'),
        ('ROWHEIGHT',  (0,0),(-1,0), 26),
        ('ROWHEIGHT',  (0,1),(-1,1), 45),
        ('BOX',        (0,0),(-1,-1),1.5, VERT),
    ]))
    el.append(total_t)

    el.append(Spacer(1, 0.5*cm))
    el.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    el.append(Paragraph(
        f"Document généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} — Légal Connect — Confidentiel",
        s_footer
    ))
    doc.build(el)
    return response
