from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import FileResponse, Http404, HttpResponse
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from datetime import date, timedelta
import os
import json

from .models import Document, VersionDocument, CommentaireDocument, CATEGORIES
from .forms import DocumentForm


# ── Helpers ──────────────────────────────────────────────────────────────────

def _acces_document(user, pk, proprietaire_only=False):
    """Retourne le document si l'utilisateur y a accès, sinon lève Http404."""
    profil = getattr(user, 'profil', None)
    role   = profil.role if profil else 'commercant'

    if proprietaire_only:
        return get_object_or_404(Document, pk=pk, proprietaire=user)

    if role in ['admin', 'institution']:
        return get_object_or_404(Document, pk=pk)
    elif role == 'expert':
        doc = get_object_or_404(Document, pk=pk)
        if doc.proprietaire != user and not (doc.statut == 'partage' and
            (doc.expert == user or user in doc.experts_partage.all())):
            raise Http404
        return doc
    else:
        return get_object_or_404(Document, pk=pk, proprietaire=user)


def _experts():
    return User.objects.filter(profil__role='expert', is_active=True).select_related('profil')


# ── Liste ─────────────────────────────────────────────────────────────────────

@login_required
def liste_documents(request):
    user   = request.user
    profil = getattr(user, 'profil', None)
    role   = profil.role if profil else 'commercant'

    if role in ['admin', 'institution']:
        documents = Document.objects.all()
    elif role == 'expert':
        documents = Document.objects.filter(
            Q(proprietaire=user) |
            Q(expert=user, statut='partage') |
            Q(experts_partage=user, statut='partage')
        ).distinct()
    else:
        documents = Document.objects.filter(proprietaire=user)

    # Filtres
    categorie  = request.GET.get('categorie', '')
    statut     = request.GET.get('statut', '')
    recherche  = request.GET.get('q', '')
    tri        = request.GET.get('tri', '-created_at')
    expire     = request.GET.get('expire', '')

    if categorie:
        documents = documents.filter(categorie=categorie)
    if statut:
        documents = documents.filter(statut=statut)
    if recherche:
        documents = documents.filter(
            Q(titre__icontains=recherche) |
            Q(description__icontains=recherche) |
            Q(mots_cles__icontains=recherche) |
            Q(numero_reference__icontains=recherche)
        )
    if expire == '1':
        documents = documents.filter(date_expiration__lt=date.today())
    elif expire == 'soon':
        documents = documents.filter(
            date_expiration__gte=date.today(),
            date_expiration__lte=date.today() + timedelta(days=30)
        )

    tris_valides = ['titre', '-titre', 'created_at', '-created_at', 'taille', '-taille', 'date_expiration']
    if tri in tris_valides:
        documents = documents.order_by(tri)

    # Stats globales
    total      = documents.count()
    total_size = documents.aggregate(s=Sum('taille'))['s'] or 0
    if total_size < 1024 * 1024:
        total_size_str = f"{total_size // 1024} Ko"
    else:
        total_size_str = f"{total_size / (1024*1024):.1f} Mo"

    docs_expires   = documents.filter(date_expiration__lt=date.today()).count()
    docs_bientot   = documents.filter(
        date_expiration__gte=date.today(),
        date_expiration__lte=date.today() + timedelta(days=30)
    ).count()

    # Pagination
    paginator = Paginator(documents, 12)
    page      = request.GET.get('page', 1)
    page_obj  = paginator.get_page(page)

    context = {
        'page_obj'     : page_obj,
        'documents'    : page_obj,
        'categories'   : CATEGORIES,
        'cat_active'   : categorie,
        'statut_actif' : statut,
        'recherche'    : recherche,
        'tri'          : tri,
        'expire'       : expire,
        'total'        : total,
        'total_size'   : total_size_str,
        'docs_expires' : docs_expires,
        'docs_bientot' : docs_bientot,
        'vue'          : request.GET.get('vue', 'cartes'),
    }
    return render(request, 'documents/liste.html', context)


# ── Upload ────────────────────────────────────────────────────────────────────

@login_required
def upload_document(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc              = form.save(commit=False)
            doc.proprietaire = request.user
            if request.FILES.get('fichier'):
                doc.taille = request.FILES['fichier'].size
            doc.save()
            messages.success(request, f'Document « {doc.titre} » uploadé avec succès !')
            return redirect('documents:detail', pk=doc.pk)
    else:
        form = DocumentForm()
    return render(request, 'documents/upload.html', {'form': form})


# ── Détail ────────────────────────────────────────────────────────────────────

@login_required
def detail_document(request, pk):
    doc          = _acces_document(request.user, pk)
    commentaires = doc.commentaires.select_related('auteur').all()
    versions     = doc.versions.select_related('cree_par').all()

    context = {
        'doc'         : doc,
        'experts'     : _experts(),
        'commentaires': commentaires,
        'versions'    : versions,
        'est_proprio' : doc.proprietaire == request.user,
    }
    return render(request, 'documents/detail.html', context)


# ── Modifier ──────────────────────────────────────────────────────────────────

@login_required
def modifier_document(request, pk):
    doc = _acces_document(request.user, pk, proprietaire_only=True)

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=doc)
        if form.is_valid():
            note_version = request.POST.get('note_version', '')
            ancien_fichier = doc.fichier

            # Sauvegarder l'ancienne version avant modification
            if request.FILES.get('fichier') and ancien_fichier:
                nb_versions = doc.versions.count()
                VersionDocument.objects.create(
                    document   = doc,
                    fichier    = ancien_fichier,
                    taille     = doc.taille,
                    note       = note_version or f'Version {nb_versions + 1}',
                    cree_par   = request.user,
                    numero     = nb_versions + 1,
                )

            doc = form.save(commit=False)
            if request.FILES.get('fichier'):
                doc.taille = request.FILES['fichier'].size
            doc.save()
            messages.success(request, f'Document « {doc.titre} » mis à jour avec succès !')
            return redirect('documents:detail', pk=doc.pk)
    else:
        form = DocumentForm(instance=doc)

    return render(request, 'documents/modifier.html', {'form': form, 'doc': doc})


# ── Supprimer ─────────────────────────────────────────────────────────────────

@login_required
def supprimer_document(request, pk):
    doc = _acces_document(request.user, pk, proprietaire_only=True)

    if request.method == 'POST':
        titre = doc.titre
        # Supprimer fichier principal
        try:
            if doc.fichier and os.path.isfile(doc.fichier.path):
                os.remove(doc.fichier.path)
        except OSError:
            pass  # fichier verrouille/deja absent : on supprime quand meme la fiche
        # Supprimer fichiers de versions
        for v in doc.versions.all():
            try:
                if v.fichier and os.path.isfile(v.fichier.path):
                    os.remove(v.fichier.path)
            except Exception:
                pass
        doc.delete()
        messages.success(request, f'Document « {titre} » supprimé définitivement.')
        return redirect('documents:liste')

    return render(request, 'documents/supprimer.html', {'doc': doc})


# ── Téléchargement sécurisé ───────────────────────────────────────────────────

@login_required
def telecharger_document(request, pk):
    doc = _acces_document(request.user, pk)

    if not doc.fichier or not os.path.isfile(doc.fichier.path):
        messages.error(request, 'Fichier introuvable.')
        return redirect('documents:detail', pk=pk)

    doc.nb_telechargements += 1
    doc.save(update_fields=['nb_telechargements'])

    response = FileResponse(open(doc.fichier.path, 'rb'))
    response['Content-Disposition'] = (
        f'attachment; filename="{os.path.basename(doc.fichier.name)}"'
    )
    return response


@login_required
def telecharger_version(request, pk, version_pk):
    doc     = _acces_document(request.user, pk)
    version = get_object_or_404(VersionDocument, pk=version_pk, document=doc)

    if not version.fichier or not os.path.isfile(version.fichier.path):
        messages.error(request, 'Fichier de version introuvable.')
        return redirect('documents:detail', pk=pk)

    response = FileResponse(open(version.fichier.path, 'rb'))
    response['Content-Disposition'] = (
        f'attachment; filename="v{version.numero}_{os.path.basename(version.fichier.name)}"'
    )
    return response


# ── Partage ───────────────────────────────────────────────────────────────────

@login_required
def partager_document(request, pk):
    doc = _acces_document(request.user, pk, proprietaire_only=True)

    if request.method == 'POST':
        expert_ids = request.POST.getlist('expert_ids')
        if expert_ids:
            experts = User.objects.filter(pk__in=expert_ids)
            doc.experts_partage.set(experts)
            # Rétro-compat : garder le premier comme doc.expert
            doc.expert = experts.first() if experts.exists() else None
            doc.statut = 'partage' if experts.exists() else 'prive'
            doc.save()
            noms = ', '.join(e.get_full_name() or e.username for e in experts)
            messages.success(request, f'Document partagé avec : {noms}')
        else:
            # Révoquer tous les partages
            doc.experts_partage.clear()
            doc.expert = None
            doc.statut = 'prive'
            doc.save()
            messages.info(request, 'Partage révoqué — document repassé en privé.')

    return redirect('documents:detail', pk=pk)


# ── Archiver / Désarchiver ────────────────────────────────────────────────────

@login_required
def archiver_document(request, pk):
    doc        = _acces_document(request.user, pk, proprietaire_only=True)
    doc.statut = 'archive'
    doc.save()
    messages.success(request, f'Document « {doc.titre} » archivé.')
    return redirect('documents:liste')


@login_required
def desarchiver_document(request, pk):
    doc        = _acces_document(request.user, pk, proprietaire_only=True)
    doc.statut = 'prive'
    doc.save()
    messages.success(request, f'Document « {doc.titre} » restauré.')
    return redirect('documents:detail', pk=pk)


# ── Commentaires ──────────────────────────────────────────────────────────────

@login_required
def ajouter_commentaire(request, pk):
    doc = _acces_document(request.user, pk)

    if request.method == 'POST':
        texte = request.POST.get('texte', '').strip()
        if texte:
            est_expert = (
                hasattr(request.user, 'profil') and
                request.user.profil.role == 'expert'
            )
            CommentaireDocument.objects.create(
                document   = doc,
                auteur     = request.user,
                texte      = texte,
                est_expert = est_expert,
            )
            messages.success(request, 'Commentaire ajouté.')
        else:
            messages.error(request, 'Le commentaire ne peut pas être vide.')

    return redirect('documents:detail', pk=pk)


# ── Statistiques ──────────────────────────────────────────────────────────────

@login_required
def statistiques_documents(request):
    user   = request.user
    profil = getattr(user, 'profil', None)
    role   = profil.role if profil else 'commercant'

    if role in ['admin', 'institution']:
        docs = Document.objects.all()
    elif role == 'expert':
        docs = Document.objects.filter(
            Q(proprietaire=user) |
            Q(expert=user, statut='partage') |
            Q(experts_partage=user, statut='partage')
        ).distinct()
    else:
        docs = Document.objects.filter(proprietaire=user)

    aujourd_hui = date.today()

    # Stats globales
    total          = docs.count()
    total_size     = docs.aggregate(s=Sum('taille'))['s'] or 0
    total_dl       = docs.aggregate(s=Sum('nb_telechargements'))['s'] or 0
    docs_partages  = docs.filter(statut='partage').count()
    docs_archives  = docs.filter(statut='archive').count()
    docs_expires   = docs.filter(date_expiration__lt=aujourd_hui).count()
    docs_bientot   = docs.filter(
        date_expiration__gte=aujourd_hui,
        date_expiration__lte=aujourd_hui + timedelta(days=30)
    ).count()

    # Par catégorie
    par_cat = []
    for code, label in CATEGORIES:
        nb    = docs.filter(categorie=code).count()
        taille = docs.filter(categorie=code).aggregate(s=Sum('taille'))['s'] or 0
        if nb > 0:
            par_cat.append({'code': code, 'label': label, 'nb': nb, 'taille': taille})

    # Par statut
    par_statut = {
        'prive'      : docs.filter(statut='prive').count(),
        'partage'    : docs.filter(statut='partage').count(),
        'en_revision': docs.filter(statut='en_revision').count(),
        'signe'      : docs.filter(statut='signe').count(),
        'archive'    : docs.filter(statut='archive').count(),
    }

    # Évolution mensuelle (6 derniers mois)
    evol_labels, evol_data = [], []
    for i in range(5, -1, -1):
        m = aujourd_hui.month - i
        y = aujourd_hui.year
        while m <= 0:
            m += 12; y -= 1
        nb_m = docs.filter(created_at__year=y, created_at__month=m).count()
        from calendar import month_abbr
        evol_labels.append(f"{month_abbr[m]} {y}")
        evol_data.append(nb_m)

    # Taille formatée
    if total_size < 1024 * 1024:
        total_size_str = f"{total_size // 1024} Ko"
    elif total_size < 1024 ** 3:
        total_size_str = f"{total_size / (1024*1024):.1f} Mo"
    else:
        total_size_str = f"{total_size / (1024**3):.2f} Go"

    context = {
        'total'          : total,
        'total_size'     : total_size_str,
        'total_dl'       : total_dl,
        'docs_partages'  : docs_partages,
        'docs_archives'  : docs_archives,
        'docs_expires'   : docs_expires,
        'docs_bientot'   : docs_bientot,
        'par_cat'        : par_cat,
        'par_statut'     : par_statut,
        'chart_cat_labels': json.dumps([c['label'] for c in par_cat]),
        'chart_cat_data'  : json.dumps([c['nb'] for c in par_cat]),
        'chart_evol_labels': json.dumps(evol_labels),
        'chart_evol_data'  : json.dumps(evol_data),
        'chart_statut_labels': json.dumps(list(par_statut.keys())),
        'chart_statut_data'  : json.dumps(list(par_statut.values())),
    }
    return render(request, 'documents/statistiques.html', context)


# ── Export Excel ──────────────────────────────────────────────────────────────

@login_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user   = request.user
    profil = getattr(user, 'profil', None)
    role   = profil.role if profil else 'commercant'

    if role in ['admin', 'institution']:
        docs = Document.objects.all()
    else:
        docs = Document.objects.filter(proprietaire=user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mes Documents"

    BLEU  = "1B4F72"

    entete_font  = Font(bold=True, color="FFFFFF", size=11)
    entete_fill  = PatternFill("solid", fgColor=BLEU)
    entete_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    titre_font   = Font(bold=True, color=BLEU, size=14)

    ws.merge_cells('A1:H1')
    ws['A1'] = "LÉGAL CONNECT — Mes Documents"
    ws['A1'].font = titre_font
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Exporté le {date.today().strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(italic=True, color="777777")
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.append([])

    colonnes = ['Titre', 'Catégorie', 'Statut', 'Type', 'Taille', 'Référence', 'Expiration', 'Date upload']
    ws.append(colonnes)
    row_h = ws.max_row
    for i, _ in enumerate(colonnes, 1):
        c = ws.cell(row=row_h, column=i)
        c.font = entete_font; c.fill = entete_fill
        c.alignment = entete_align; c.border = thin_border

    alt_fill = PatternFill("solid", fgColor="F0F7FF")
    for idx, d in enumerate(docs):
        row = [
            d.titre,
            d.get_categorie_display(),
            d.get_statut_display(),
            d.extension.upper(),
            d.taille_lisible,
            d.numero_reference or '—',
            d.date_expiration.strftime('%d/%m/%Y') if d.date_expiration else '—',
            d.created_at.strftime('%d/%m/%Y'),
        ]
        ws.append(row)
        r = ws.max_row
        fill = alt_fill if idx % 2 == 0 else None
        for col in range(1, len(colonnes) + 1):
            c = ws.cell(row=r, column=col)
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
            if fill:
                c.fill = fill

    largeurs = [35, 22, 18, 8, 10, 18, 14, 14]
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row_h].height = 28

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="documents_{date.today().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response
